"""이체결과내역 (BZ뱅크 grid_exceldata) import — transfer_memo 보강.

신규 transactions 생성 X — 기존 거래 (Codef sync / Excel 업로드) 와
date + amount + counterparty 매칭 후 transfer_memo 만 update.

사용자가 이체 시 입력한 메모 ('주정차과태료', '본사임대료', '운영자금 상환'
등) 가 통장 거래내역에는 누락 — 이 메모로 매핑 정확도 향상.
"""

import io
from dataclasses import dataclass

import openpyxl
from psycopg2.extensions import connection as PgConnection


@dataclass
class TransferRow:
    date: str           # YYYY-MM-DD
    amount: float
    payee: str          # 수취인성명
    memo: str           # 거래메모
    payee_display: str  # 입금통장표시내용 (보통 payee 의 일부)


@dataclass
class ImportResult:
    total: int
    matched: int
    ambiguous: int
    unmatched: int
    sample_matched: list[dict]
    ambiguous_rows: list[dict]
    unmatched_rows: list[dict]


REQUIRED_HEADERS = ("거래일시", "수취인성명", "거래메모")


def parse_transfer_history(file_bytes: bytes) -> list[TransferRow]:
    """xlsx → list[TransferRow]. BZ뱅크 grid_exceldata 형식.

    헤더 컬럼: No / 전체선택 / 이체구분 / 거래일시 / 처리결과 / 출금계좌 /
    입금은행 / 입금계좌 / 수취인성명 / 처리금액 / 입금금액 / 수수료 /
    오류금액 / 등록일자 / 오류코드 / 거래메모 / 입금인코드 /
    출금통장표시내용 / 입금통장표시내용 / 수취인수수료부담여부
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    sh = wb[wb.sheetnames[0]]

    # 헤더 row 찾기 (1행 또는 2행)
    header_row = None
    for r in (1, 2):
        if r > sh.max_row:
            break
        cells = [str(sh.cell(r, c).value or "").strip() for c in range(1, sh.max_column + 1)]
        if all(h in cells for h in REQUIRED_HEADERS):
            header_row = r
            break
    if header_row is None:
        raise ValueError("이체결과내역 헤더(거래일시/수취인성명/거래메모)를 찾을 수 없습니다")

    headers = [str(sh.cell(header_row, c).value or "").strip() for c in range(1, sh.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers)}  # 1-indexed
    col_dt = col["거래일시"]
    col_payee = col["수취인성명"]
    col_memo = col["거래메모"]
    col_amt = col.get("입금금액") or col.get("처리금액")
    col_payee_disp = col.get("입금통장표시내용")
    if not col_amt:
        raise ValueError("입금금액 또는 처리금액 컬럼이 없습니다")

    rows = []
    for r in range(header_row + 1, sh.max_row + 1):
        dt_str = str(sh.cell(r, col_dt).value or "")
        if "." not in dt_str:
            continue
        date_part = dt_str.split(" ")[0].replace(".", "-")
        amount_val = sh.cell(r, col_amt).value
        if amount_val is None or amount_val == "":
            continue
        try:
            amount = float(amount_val)
        except (TypeError, ValueError):
            continue
        payee = str(sh.cell(r, col_payee).value or "").strip()
        memo = str(sh.cell(r, col_memo).value or "").strip()
        payee_display = str(sh.cell(r, col_payee_disp).value or "").strip() if col_payee_disp else ""
        rows.append(TransferRow(date=date_part, amount=amount, payee=payee,
                                memo=memo, payee_display=payee_display))
    return rows


def apply_transfer_memos(
    conn: PgConnection,
    entity_id: int,
    rows: list[TransferRow],
    overwrite: bool = False,
) -> ImportResult:
    """기존 transactions 매칭 + transfer_memo update.

    매칭 키: date + amount + type='out'.
    여러 후보 시 counterparty 부분 일치로 disambiguate.
    overwrite=False (기본): 기존 transfer_memo 가 있으면 skip.
    overwrite=True: 기존 값 덮어쓰기.
    """
    cur = conn.cursor()
    cur.execute("SET search_path TO financeone, public")

    matched = 0
    ambiguous_rows = []
    unmatched_rows = []
    sample_matched = []

    for row in rows:
        if not row.memo:
            continue

        cur.execute(
            """
            SELECT id, counterparty, transfer_memo
            FROM transactions
            WHERE entity_id = %s AND date = %s AND amount = %s AND type = 'out'
              AND is_duplicate = false AND (is_cancel IS NOT TRUE)
            """,
            [entity_id, row.date, row.amount],
        )
        candidates = cur.fetchall()

        if len(candidates) == 0:
            unmatched_rows.append({
                "date": row.date, "amount": row.amount,
                "payee": row.payee, "memo": row.memo,
                "reason": "no match",
            })
            continue

        # Disambiguate by counterparty / payee_display
        if len(candidates) > 1:
            payee_keys = [k for k in (row.payee_display, row.payee) if k]
            picked = None
            for cand in candidates:
                cp = cand[1] or ""
                for key in payee_keys:
                    if key and (key[:6] in cp or cp[:6] in key):
                        picked = cand
                        break
                if picked:
                    break
            if not picked:
                ambiguous_rows.append({
                    "date": row.date, "amount": row.amount,
                    "payee": row.payee, "memo": row.memo,
                    "candidate_count": len(candidates),
                })
                continue
            tx_id, cp, existing_memo = picked
        else:
            tx_id, cp, existing_memo = candidates[0]

        if existing_memo and not overwrite:
            # 이미 메모 있음 — skip
            continue

        cur.execute(
            "UPDATE transactions SET transfer_memo = %s WHERE id = %s",
            [row.memo, tx_id],
        )
        matched += 1
        if len(sample_matched) < 10:
            sample_matched.append({
                "tx_id": tx_id, "date": row.date,
                "amount": row.amount, "memo": row.memo,
                "counterparty": cp,
            })

    cur.close()

    return ImportResult(
        total=len(rows),
        matched=matched,
        ambiguous=len(ambiguous_rows),
        unmatched=len(unmatched_rows),
        sample_matched=sample_matched,
        ambiguous_rows=ambiguous_rows,
        unmatched_rows=unmatched_rows,
    )

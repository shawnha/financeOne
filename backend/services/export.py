"""재무제표 Excel/PDF Export"""

import io
from psycopg2.extensions import connection as PgConnection
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter


STATEMENT_LABELS = {
    "balance_sheet": "재무상태표",
    "income_statement": "P&L",
    "cash_flow": "현금흐름표",
    "trial_balance": "합계잔액시산표",
    "deficit_treatment": "결손금처리계산서",
}

# 거래내역 export 시 source_type 분류 — cashflow_service 와 동일하게 유지
EXPORT_BANK_SOURCES = (
    "woori_bank", "codef_woori_bank",
    "ibk_bank", "codef_ibk_bank",
    "shinhan_bank", "codef_shinhan_bank",
    "mercury_api", "manual",
)
EXPORT_CARD_SOURCES = (
    "lotte_card", "woori_card", "shinhan_card",
    "codef_lotte_card", "codef_woori_card", "codef_shinhan_card",
)

# source_type → 한글 라벨 (Excel 출처 컬럼에 표시). codef_ 접두사 제거 + 한글화.
SOURCE_LABELS = {
    "woori_bank": "우리은행",
    "codef_woori_bank": "우리은행",
    "ibk_bank": "IBK기업은행",
    "codef_ibk_bank": "IBK기업은행",
    "shinhan_bank": "신한은행",
    "codef_shinhan_bank": "신한은행",
    "mercury_api": "Mercury",
    "manual": "수기입력",
    "woori_card": "우리카드",
    "codef_woori_card": "우리카드",
    "lotte_card": "롯데카드",
    "codef_lotte_card": "롯데카드",
    "shinhan_card": "신한카드",
    "codef_shinhan_card": "신한카드",
}

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
header_font = Font(name="맑은 고딕", bold=True, size=11, color="FFFFFF")
section_font = Font(name="맑은 고딕", bold=True, size=10)
normal_font = Font(name="맑은 고딕", size=10)
amount_font = Font(name="Consolas", size=10)


def export_statement_excel(
    conn: PgConnection,
    statement_id: int,
    statement_type: str | None = None,
) -> bytes:
    """재무제표를 Excel로 Export.

    Args:
        statement_type: None이면 전체, 지정하면 해당 유형만

    Returns:
        Excel 파일 바이트
    """
    cur = conn.cursor()

    # 헤더 정보
    cur.execute(
        """
        SELECT fs.fiscal_year, fs.start_month, fs.end_month,
               e.name AS entity_name
        FROM financial_statements fs
        LEFT JOIN entities e ON fs.entity_id = e.id
        WHERE fs.id = %s
        """,
        [statement_id],
    )
    header = cur.fetchone()
    if not header:
        cur.close()
        raise ValueError(f"Statement {statement_id} not found")

    fiscal_year, start_month, end_month, entity_name = header

    # 라인 아이템
    type_filter = "AND li.statement_type = %s" if statement_type else ""
    params = [statement_id]
    if statement_type:
        params.append(statement_type)

    cur.execute(
        f"""
        SELECT li.statement_type, li.account_code, li.label,
               li.is_section_header,
               COALESCE(li.manual_amount, li.auto_amount) AS amount,
               COALESCE(li.manual_debit, li.auto_debit) AS debit,
               COALESCE(li.manual_credit, li.auto_credit) AS credit
        FROM financial_statement_line_items li
        WHERE li.statement_id = %s {type_filter}
        ORDER BY li.statement_type, li.sort_order
        """,
        params,
    )
    rows = cur.fetchall()
    cur.close()

    # Group by statement type
    grouped: dict[str, list] = {}
    for row in rows:
        st = row[0]
        if st not in grouped:
            grouped[st] = []
        grouped[st].append(row)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    for st_type, items in grouped.items():
        label = STATEMENT_LABELS.get(st_type, st_type)
        ws = wb.create_sheet(title=label[:31])  # Excel sheet name limit

        # Title
        ws.merge_cells("A1:D1")
        ws["A1"] = f"{entity_name} — {label}"
        ws["A1"].font = Font(name="맑은 고딕", bold=True, size=14)
        ws["A2"] = f"{fiscal_year}년 {start_month}월~{end_month}월"
        ws["A2"].font = Font(name="맑은 고딕", size=10, color="888888")

        is_tb = st_type == "trial_balance"

        # Headers
        row_num = 4
        ws.cell(row=row_num, column=1, value="계정과목").font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        ws.cell(row=row_num, column=1).border = thin_border

        if is_tb:
            ws.cell(row=row_num, column=2, value="차변").font = header_font
            ws.cell(row=row_num, column=2).fill = header_fill
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="right")
            ws.cell(row=row_num, column=3, value="대변").font = header_font
            ws.cell(row=row_num, column=3).fill = header_fill
            ws.cell(row=row_num, column=3).border = thin_border
            ws.cell(row=row_num, column=3).alignment = Alignment(horizontal="right")
        else:
            ws.cell(row=row_num, column=2, value="금액").font = header_font
            ws.cell(row=row_num, column=2).fill = header_fill
            ws.cell(row=row_num, column=2).border = thin_border
            ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="right")

        # Data rows
        for item in items:
            row_num += 1
            _, account_code, item_label, is_header, amount, debit, credit = item

            cell_label = ws.cell(row=row_num, column=1, value=item_label)
            cell_label.font = section_font if is_header else normal_font
            cell_label.border = thin_border

            if is_tb:
                cell_d = ws.cell(row=row_num, column=2, value=float(debit) if debit else None)
                cell_d.font = amount_font
                cell_d.number_format = '#,##0'
                cell_d.alignment = Alignment(horizontal="right")
                cell_d.border = thin_border

                cell_c = ws.cell(row=row_num, column=3, value=float(credit) if credit else None)
                cell_c.font = amount_font
                cell_c.number_format = '#,##0'
                cell_c.alignment = Alignment(horizontal="right")
                cell_c.border = thin_border
            else:
                cell_a = ws.cell(row=row_num, column=2, value=float(amount) if amount else None)
                cell_a.font = amount_font
                cell_a.number_format = '#,##0'
                cell_a.alignment = Alignment(horizontal="right")
                cell_a.border = thin_border

        # Column widths
        ws.column_dimensions["A"].width = 40
        ws.column_dimensions["B"].width = 20
        if is_tb:
            ws.column_dimensions["C"].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def export_transactions_excel(
    conn: PgConnection,
    entity_id: int,
    year: int,
    month: int,
    kind: str = "all",
    where_clause: str | None = None,
    params: list | None = None,
    need_join_for_search: bool = False,
    filtered: bool = False,
) -> bytes:
    """거래내역 Excel Export (회계법인 전달용).

    월별 단일 sheet — 한 entity 의 한 달 거래.
    kind: 'all' (전체) | 'bank' (은행만) | 'card' (카드만)
    컬럼: 날짜 / 시각 / 출처 / 회원 / 적요 / 거래처 / 입금 / 출금 /
          내부계정 / 표준계정코드 / 표준계정명 / 메모 / 취소

    where_clause / params: list_transactions 와 공유되는 필터 WHERE.
        None 이면 entity_id + 월간 기본 WHERE 만 적용.
    need_join_for_search: WHERE 가 ia.name 을 참조하는지 여부 (export 는 항상 ia JOIN 하므로 무시 가능).
    filtered: 사용자 필터가 걸려 있어 부분만 export 된 경우 (제목/시트명에 ' (필터)' 부착).
    """
    import calendar

    if where_clause is None:
        last_day = calendar.monthrange(year, month)[1]
        start_d = f"{year:04d}-{month:02d}-01"
        end_d = f"{year:04d}-{month:02d}-{last_day:02d}"
        where_clause = "t.entity_id = %s AND t.date >= %s AND t.date <= %s"
        params = [entity_id, start_d, end_d]
    else:
        params = list(params or [])

    if kind == "bank":
        where_clause = f"({where_clause}) AND t.source_type = ANY(%s)"
        params.append(list(EXPORT_BANK_SOURCES))
    elif kind == "card":
        where_clause = f"({where_clause}) AND t.source_type = ANY(%s)"
        params.append(list(EXPORT_CARD_SOURCES))

    cur = conn.cursor()
    cur.execute(
        f"""
        SELECT t.date, t.time, t.source_type, m.name AS member_name,
               t.description, t.counterparty,
               CASE WHEN t.type = 'in' THEN t.amount ELSE NULL END AS in_amt,
               CASE WHEN t.type = 'out' THEN t.amount ELSE NULL END AS out_amt,
               ia.name AS ia_name,
               sa.code AS sa_code, sa.name AS sa_name,
               t.note, t.is_cancel
        FROM transactions t
        LEFT JOIN members m ON m.id = t.member_id
        LEFT JOIN internal_accounts ia ON ia.id = t.internal_account_id
        LEFT JOIN standard_accounts sa ON sa.id = t.standard_account_id
        WHERE {where_clause}
        ORDER BY t.date, t.time NULLS LAST, t.id
        """,
        params,
    )
    rows = cur.fetchall()

    cur.execute("SELECT name FROM entities WHERE id = %s", [entity_id])
    entity_row = cur.fetchone()
    entity_name = entity_row[0] if entity_row else f"entity_{entity_id}"
    cur.close()

    kind_label = {"bank": "은행", "card": "카드"}.get(kind, "")
    filter_label = " (필터)" if filtered else ""
    sheet_title = f"{year}-{month:02d}" + (f" {kind_label}" if kind_label else "") + filter_label
    title_text = (
        f"{entity_name} 거래내역"
        + (f" — {kind_label}" if kind_label else "")
        + f" ({year}-{month:02d})"
        + filter_label
    )

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_title

    title_cell = ws.cell(row=1, column=1, value=title_text)
    title_cell.font = Font(name="맑은 고딕", bold=True, size=14)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)

    headers = ["날짜", "시각", "출처", "회원", "적요", "거래처", "입금", "출금",
               "내부계정", "표준계정코드", "표준계정명", "메모", "취소"]
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border

    in_sum = 0
    out_sum = 0
    for i, r in enumerate(rows, start=4):
        date, t_time, src, member, desc, cp, in_amt, out_amt, ia_name, sa_code, sa_name, note, is_cancel = r
        # 시간 포맷팅 (HHMMSS → HH:MM:SS)
        time_fmt = ""
        if t_time and len(t_time) >= 6:
            time_fmt = f"{t_time[:2]}:{t_time[2:4]}:{t_time[4:6]}"

        ws.cell(row=i, column=1, value=date).font = normal_font
        ws.cell(row=i, column=2, value=time_fmt).font = normal_font
        ws.cell(row=i, column=3, value=SOURCE_LABELS.get(src, src) or "").font = normal_font
        ws.cell(row=i, column=4, value=member or "").font = normal_font
        ws.cell(row=i, column=5, value=desc or "").font = normal_font
        ws.cell(row=i, column=6, value=cp or "").font = normal_font

        in_cell = ws.cell(row=i, column=7, value=float(in_amt) if in_amt else None)
        in_cell.font = amount_font
        in_cell.number_format = "#,##0"
        in_cell.alignment = Alignment(horizontal="right")
        if in_amt:
            in_sum += float(in_amt)

        out_cell = ws.cell(row=i, column=8, value=float(out_amt) if out_amt else None)
        out_cell.font = amount_font
        out_cell.number_format = "#,##0"
        out_cell.alignment = Alignment(horizontal="right")
        if out_amt:
            out_sum += float(out_amt)

        ws.cell(row=i, column=9, value=ia_name or "").font = normal_font
        ws.cell(row=i, column=10, value=sa_code or "").font = normal_font
        ws.cell(row=i, column=11, value=sa_name or "").font = normal_font
        ws.cell(row=i, column=12, value=note or "").font = normal_font
        ws.cell(row=i, column=13, value="취소" if is_cancel else "").font = normal_font

    # 합계 row
    total_row = len(rows) + 4
    ws.cell(row=total_row, column=6, value="합계").font = section_font
    c_in = ws.cell(row=total_row, column=7, value=in_sum)
    c_in.font = section_font
    c_in.number_format = "#,##0"
    c_in.alignment = Alignment(horizontal="right")
    c_out = ws.cell(row=total_row, column=8, value=out_sum)
    c_out.font = section_font
    c_out.number_format = "#,##0"
    c_out.alignment = Alignment(horizontal="right")

    widths = [12, 10, 18, 12, 28, 28, 14, 14, 16, 12, 18, 28, 6]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A4"

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()

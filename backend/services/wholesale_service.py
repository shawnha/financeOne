"""도매 매출관리/매입관리 xlsx import.

매출관리(48 col) — col 41/42 매입가가 매출원가 단가 (COGS).
매입관리(40 col) — 매입 단위 row.
합계 row 자동 skip (col 7 매출구분/매입구분 비어있는 row).
"""

import io
import json
from dataclasses import dataclass
from datetime import date, datetime
from typing import Optional

import openpyxl
from psycopg2.extensions import connection as PgConnection


@dataclass
class ImportResult:
    total_rows: int
    inserted: int
    duplicates: int
    skipped: int
    sample: list[dict]
    errors: list[str]
    alerts: dict | None = None


def compute_sales_alerts(rows: list[dict]) -> dict:
    """매출관리 row 들에서 회계 이상 패턴 3종 감지.

    1. cogs_book_vs_real_diff: 매입가(장부) ≠ 매입가(실) — 가격 변동 추적
    2. negative_margin: 매출액 < 매출원가 — 손실 판매 (loss leader / 재고 처분)
    3. missing_cogs: 매입가(장부) 누락 — 매출원가 미반영
    """
    diff_count = 0
    diff_total = 0.0
    diff_examples: list[dict] = []

    neg_rows: list[dict] = []
    missing_rows: list[dict] = []

    for r in rows:
        qty = r.get("quantity") or 0
        total = r.get("total_amount") or 0
        cogs_book = r.get("cogs_unit_price")
        cogs_real = r.get("cogs_real_unit_price")

        if cogs_book is None or cogs_book == 0:
            if len(missing_rows) < 20:
                missing_rows.append({
                    "date": str(r.get("sales_date")),
                    "payee": r.get("payee_name"),
                    "product": (r.get("product_name") or "")[:40],
                    "qty": qty,
                    "total": total,
                })
            continue

        if cogs_real is not None and cogs_real != 0 and abs(cogs_book - cogs_real) > 0.001:
            diff_count += 1
            row_diff = abs(cogs_book - cogs_real) * qty
            diff_total += row_diff
            if len(diff_examples) < 10:
                diff_examples.append({
                    "date": str(r.get("sales_date")),
                    "payee": r.get("payee_name"),
                    "product": (r.get("product_name") or "")[:40],
                    "qty": qty,
                    "cogs_book": cogs_book,
                    "cogs_real": cogs_real,
                    "diff": row_diff,
                })

        cogs_total = qty * cogs_book
        margin = total - cogs_total
        if margin < 0 and total > 0:
            if len(neg_rows) < 20:
                neg_rows.append({
                    "date": str(r.get("sales_date")),
                    "payee": r.get("payee_name"),
                    "product": (r.get("product_name") or "")[:40],
                    "qty": qty,
                    "total": total,
                    "cogs_total": cogs_total,
                    "margin": margin,
                })

    return {
        "cogs_book_vs_real_diff": {
            "count": diff_count,
            "total_diff": round(diff_total, 2),
            "examples": diff_examples,
        },
        "negative_margin": {
            "count": len(neg_rows),
            "rows": neg_rows,
        },
        "missing_cogs": {
            "count": len(missing_rows),
            "rows": missing_rows,
        },
    }


def compute_purchases_alerts(rows: list[dict]) -> dict:
    """매입관리 row 들에서 회계 이상 패턴 감지.

    1. unit_price_book_vs_real_diff: 단가 장부 vs 실 차이
    2. missing_unit_price: 단가 누락
    """
    diff_count = 0
    diff_total = 0.0
    diff_examples: list[dict] = []
    missing_rows: list[dict] = []

    for r in rows:
        qty = r.get("quantity") or 0
        unit = r.get("unit_price")
        real_unit = r.get("real_unit_price")

        if unit is None or unit == 0:
            if len(missing_rows) < 20:
                missing_rows.append({
                    "date": str(r.get("purchase_date")),
                    "payee": r.get("payee_name"),
                    "product": (r.get("product_name") or "")[:40],
                    "qty": qty,
                })
            continue

        if real_unit is not None and real_unit != 0 and abs(unit - real_unit) > 0.001:
            diff_count += 1
            row_diff = abs(unit - real_unit) * qty
            diff_total += row_diff
            if len(diff_examples) < 10:
                diff_examples.append({
                    "date": str(r.get("purchase_date")),
                    "payee": r.get("payee_name"),
                    "product": (r.get("product_name") or "")[:40],
                    "qty": qty,
                    "unit_book": unit,
                    "unit_real": real_unit,
                    "diff": row_diff,
                })

    return {
        "unit_price_book_vs_real_diff": {
            "count": diff_count,
            "total_diff": round(diff_total, 2),
            "examples": diff_examples,
        },
        "missing_unit_price": {
            "count": len(missing_rows),
            "rows": missing_rows,
        },
    }


def _to_date(val) -> Optional[date]:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    for fmt in ("%Y/%m/%d", "%Y-%m-%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _to_float(val) -> Optional[float]:
    if val is None or val == "":
        return None
    try:
        return float(val)
    except (TypeError, ValueError):
        return None


def _read_sheet(file_bytes: bytes):
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
    return wb[wb.sheetnames[0]]


def parse_wholesale_sales(file_bytes: bytes) -> list[dict]:
    """매출관리 xlsx → list[dict]."""
    sh = _read_sheet(file_bytes)
    rows = []
    for r in range(3, sh.max_row + 1):
        kind = str(sh.cell(r, 7).value or "").strip()
        if kind != "매출":
            continue
        sales_date = _to_date(sh.cell(r, 2).value)
        if not sales_date:
            continue
        product = str(sh.cell(r, 9).value or "").strip()
        payee = str(sh.cell(r, 6).value or "").strip()
        if not product or not payee:
            continue

        raw = {}
        for c in range(1, sh.max_column + 1):
            val = sh.cell(r, c).value
            if val is not None and val != "":
                if isinstance(val, datetime):
                    val = val.isoformat()
                elif isinstance(val, date):
                    val = val.isoformat()
                raw[str(sh.cell(1, c).value or f"col{c}")] = val

        rows.append({
            "sales_date": sales_date,
            "document_date": _to_date(sh.cell(r, 4).value),
            "document_no": str(sh.cell(r, 5).value or "").strip() or None,
            "row_number": int(sh.cell(r, 3).value or 0) or None,
            "payee_name": payee,
            "payee_code": str(sh.cell(r, 34).value or "").strip() or None,
            "real_payee_name": str(sh.cell(r, 35).value or "").strip() or None,
            "product_name": product,
            "product_spec": str(sh.cell(r, 10).value or "").strip() or None,
            "manufacturer": str(sh.cell(r, 47).value or "").strip() or None,
            "quantity": _to_float(sh.cell(r, 11).value) or 0,
            "unit_price": _to_float(sh.cell(r, 13).value),
            "discount_pct": _to_float(sh.cell(r, 14).value),
            "supply_amount": _to_float(sh.cell(r, 15).value),
            "vat": _to_float(sh.cell(r, 16).value),
            "total_amount": _to_float(sh.cell(r, 17).value) or 0,
            "real_unit_price": _to_float(sh.cell(r, 18).value),
            "real_supply_amount": _to_float(sh.cell(r, 20).value),
            "real_total_amount": _to_float(sh.cell(r, 22).value),
            "cogs_unit_price": _to_float(sh.cell(r, 41).value),
            "cogs_real_unit_price": _to_float(sh.cell(r, 42).value),
            "bank_settled": str(sh.cell(r, 38).value or "").strip().upper() == "Y",
            "sales_rep": str(sh.cell(r, 26).value or "").strip() or None,
            "note": str(sh.cell(r, 32).value or "").strip() or None,
            "raw_data": raw,
        })
    return rows


def parse_wholesale_purchases(file_bytes: bytes) -> list[dict]:
    """매입관리 xlsx → list[dict]."""
    sh = _read_sheet(file_bytes)
    rows = []
    for r in range(3, sh.max_row + 1):
        kind = str(sh.cell(r, 7).value or "").strip()
        if kind != "매입":
            continue
        purchase_date = _to_date(sh.cell(r, 2).value)
        if not purchase_date:
            continue
        product = str(sh.cell(r, 9).value or "").strip()
        payee = str(sh.cell(r, 6).value or "").strip()
        if not product or not payee:
            continue

        raw = {}
        for c in range(1, sh.max_column + 1):
            val = sh.cell(r, c).value
            if val is not None and val != "":
                if isinstance(val, datetime):
                    val = val.isoformat()
                elif isinstance(val, date):
                    val = val.isoformat()
                raw[str(sh.cell(1, c).value or f"col{c}")] = val

        rows.append({
            "purchase_date": purchase_date,
            "document_date": _to_date(sh.cell(r, 4).value),
            "document_no": str(sh.cell(r, 5).value or "").strip() or None,
            "row_number": int(sh.cell(r, 3).value or 0) or None,
            "payee_name": payee,
            "product_name": product,
            "product_spec": str(sh.cell(r, 10).value or "").strip() or None,
            "quantity": _to_float(sh.cell(r, 11).value) or 0,
            "unit_price": _to_float(sh.cell(r, 13).value),
            "supply_amount": _to_float(sh.cell(r, 15).value),
            "vat": _to_float(sh.cell(r, 16).value),
            "total_amount": _to_float(sh.cell(r, 17).value) or 0,
            "real_unit_price": _to_float(sh.cell(r, 18).value),
            "real_supply_amount": _to_float(sh.cell(r, 20).value),
            "real_total_amount": _to_float(sh.cell(r, 22).value),
            "bank_settled": str(sh.cell(r, 33).value or "").strip().upper() == "Y",
            "note": str(sh.cell(r, 29).value or "").strip() or None,
            "raw_data": raw,
        })
    return rows


def import_wholesale_sales(
    conn: PgConnection,
    entity_id: int,
    rows: list[dict],
    source_file: Optional[str] = None,
) -> ImportResult:
    cur = conn.cursor()
    cur.execute("SET search_path TO financeone, public")
    inserted = 0
    duplicates = 0
    sample = []
    errors: list[str] = []

    for row in rows:
        try:
            cur.execute(
                """
                INSERT INTO wholesale_sales (
                    entity_id, sales_date, document_date, document_no, row_number,
                    payee_name, payee_code, real_payee_name,
                    product_name, product_spec, manufacturer,
                    quantity, unit_price, discount_pct,
                    supply_amount, vat, total_amount,
                    real_unit_price, real_supply_amount, real_total_amount,
                    cogs_unit_price, cogs_real_unit_price,
                    bank_settled, sales_rep, note, raw_data, source_file
                ) VALUES (
                    %s, %s, %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s,
                    %s, %s, %s, %s::jsonb, %s
                )
                ON CONFLICT (entity_id, sales_date, document_no, row_number, product_name)
                DO NOTHING
                RETURNING id
                """,
                [
                    entity_id, row["sales_date"], row["document_date"], row["document_no"], row["row_number"],
                    row["payee_name"], row["payee_code"], row["real_payee_name"],
                    row["product_name"], row["product_spec"], row["manufacturer"],
                    row["quantity"], row["unit_price"], row["discount_pct"],
                    row["supply_amount"], row["vat"], row["total_amount"],
                    row["real_unit_price"], row["real_supply_amount"], row["real_total_amount"],
                    row["cogs_unit_price"], row["cogs_real_unit_price"],
                    row["bank_settled"], row["sales_rep"], row["note"],
                    json.dumps(row["raw_data"], ensure_ascii=False, default=str),
                    source_file,
                ],
            )
            res = cur.fetchone()
            if res:
                inserted += 1
                if len(sample) < 5:
                    sample.append({
                        "id": res[0],
                        "date": str(row["sales_date"]),
                        "payee": row["payee_name"],
                        "product": row["product_name"][:30],
                        "total": row["total_amount"],
                    })
            else:
                duplicates += 1
        except Exception as e:
            errors.append(f"row date={row.get('sales_date')} payee={row.get('payee_name')}: {e}")

    cur.close()
    return ImportResult(
        total_rows=len(rows),
        inserted=inserted,
        duplicates=duplicates,
        skipped=0,
        sample=sample,
        errors=errors[:10],
        alerts=compute_sales_alerts(rows),
    )


def import_wholesale_purchases(
    conn: PgConnection,
    entity_id: int,
    rows: list[dict],
    source_file: Optional[str] = None,
) -> ImportResult:
    cur = conn.cursor()
    cur.execute("SET search_path TO financeone, public")
    inserted = 0
    duplicates = 0
    sample = []
    errors: list[str] = []

    for row in rows:
        try:
            cur.execute(
                """
                INSERT INTO wholesale_purchases (
                    entity_id, purchase_date, document_date, document_no, row_number,
                    payee_name,
                    product_name, product_spec,
                    quantity, unit_price,
                    supply_amount, vat, total_amount,
                    real_unit_price, real_supply_amount, real_total_amount,
                    bank_settled, note, raw_data, source_file
                ) VALUES (
                    %s, %s, %s, %s, %s,
                    %s,
                    %s, %s,
                    %s, %s,
                    %s, %s, %s,
                    %s, %s, %s,
                    %s, %s, %s::jsonb, %s
                )
                ON CONFLICT (entity_id, purchase_date, document_no, row_number, product_name)
                DO NOTHING
                RETURNING id
                """,
                [
                    entity_id, row["purchase_date"], row["document_date"], row["document_no"], row["row_number"],
                    row["payee_name"],
                    row["product_name"], row["product_spec"],
                    row["quantity"], row["unit_price"],
                    row["supply_amount"], row["vat"], row["total_amount"],
                    row["real_unit_price"], row["real_supply_amount"], row["real_total_amount"],
                    row["bank_settled"], row["note"],
                    json.dumps(row["raw_data"], ensure_ascii=False, default=str),
                    source_file,
                ],
            )
            res = cur.fetchone()
            if res:
                inserted += 1
                if len(sample) < 5:
                    sample.append({
                        "id": res[0],
                        "date": str(row["purchase_date"]),
                        "payee": row["payee_name"],
                        "product": row["product_name"][:30],
                        "total": row["total_amount"],
                    })
            else:
                duplicates += 1
        except Exception as e:
            errors.append(f"row date={row.get('purchase_date')} payee={row.get('payee_name')}: {e}")

    cur.close()
    return ImportResult(
        total_rows=len(rows),
        inserted=inserted,
        duplicates=duplicates,
        skipped=0,
        sample=sample,
        errors=errors[:10],
        alerts=compute_purchases_alerts(rows),
    )

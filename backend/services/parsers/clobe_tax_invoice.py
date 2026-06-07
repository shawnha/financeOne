# 클로브(clobe.ai) 세금계산서 엑셀 파서 — 전체/매출/매입 시트를 invoices 후보 행으로 변환
"""clobe.ai 세금계산서 다운로드 엑셀 파서.

clobe 워크스페이스 > 세금계산서 > 엑셀 다운로드 양식 (고정 컬럼):
  발급일자 | 작성일자 | 매출 매입 유형 | 과세 유형 | 거래처 상호 |
  거래처 사업자등록번호 | 대표 품목 | 공급가액 | 세액 | 합계금액 |
  수정 여부 | 입금 예정일자 | 내 사업자번호

홈택스 표준 양식과 달리 '공급자/공급받는자' 분리 컬럼이 없고, 대신
'매출 매입 유형' 이 direction 을 직접 준다 → our_biz_no 추론 불필요.
승인번호 컬럼이 없어 dedup 은 (entity, direction, 작성일자, 합계금액) 자연키로 한다.
음수 행 = 수정(취소) 세금계산서.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Optional

import openpyxl


# clobe 고정 헤더 → 표준 필드
_CLOBE_HEADERS = {
    "발급일자": "issue_date_raw",
    "작성일자": "write_date",
    "매출 매입 유형": "io_type",
    "과세 유형": "tax_type",
    "거래처 상호": "counterparty",
    "거래처 사업자등록번호": "counterparty_biz_no",
    "대표 품목": "description",
    "공급가액": "amount",
    "세액": "vat",
    "합계금액": "total",
    "수정 여부": "amend",
    "입금 예정일자": "due_date",
    "내 사업자번호": "entity_biz_no",
}


def _norm_biz(s) -> Optional[str]:
    if s is None:
        return None
    digits = "".join(c for c in str(s) if c.isdigit())
    return digits or None


def _parse_date(v) -> Optional[date]:
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan"):
        return None
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 8:
        try:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            pass
    for sep in ("-", "/", "."):
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) >= 3:
                try:
                    y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return date(y, m, d)
                except (ValueError, IndexError):
                    continue
    return None


def _dec(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    s = str(v).replace(",", "").replace("₩", "").replace("원", "").strip()
    if not s or s.lower() in ("none", "nan", "-"):
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _io_to_direction(v) -> Optional[str]:
    s = str(v or "").strip()
    if "매출" in s:
        return "sales"
    if "매입" in s:
        return "purchase"
    return None


def _header_map(header_row: list) -> dict:
    """헤더 행 → {col_index: field}. clobe 헤더 정확 일치(공백 정규화)."""
    col_map = {}
    for j, cell in enumerate(header_row):
        if cell is None:
            continue
        key = str(cell).strip()
        if key in _CLOBE_HEADERS:
            col_map[j] = _CLOBE_HEADERS[key]
    return col_map


def _rows_from_sheet(ws) -> tuple[list, list]:
    """워크시트 → (header_row, data_rows). 헤더 못 찾으면 (None, [])."""
    rows = [list(r) for r in ws.iter_rows(values_only=True)]
    for i, row in enumerate(rows[:5]):
        cm = _header_map(row)
        if "amount" in cm.values() and "total" in cm.values():
            return row, rows[i + 1:]
    return None, []


def parse_clobe_tax_invoice(file_bytes: bytes) -> dict:
    """clobe 세금계산서 엑셀 → 파싱 결과.

    '전체' 시트에 데이터가 있으면 그것만, 없으면 '매출'+'매입' 시트 결합.
    direction 은 '매출 매입 유형' 컬럼 기준.

    Returns:
        {
          "parsed": [ {entity_biz_no, direction, issue_date(ISO),
                       counterparty, counterparty_biz_no, description,
                       amount, vat, total, tax_type, due_date, amend, raw}, ... ],
          "errors": [ {"sheet":..,"row":..,"message":..}, ... ],
          "stats": {"total":N,"valid":M,"errors":K,"sales":a,"purchase":b},
        }
    """
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    try:
        sheet_names = wb.sheetnames
        # '전체' 시트에 데이터가 있으면 그것만, 아니면 매출/매입 결합
        targets: list[str] = []
        if "전체" in sheet_names:
            _, data = _rows_from_sheet(wb["전체"])
            if any(r and any(c not in (None, "") for c in r) for r in data):
                targets = ["전체"]
        if not targets:
            targets = [n for n in ("매출", "매입") if n in sheet_names] or sheet_names[:1]

        parsed: list = []
        errors: list = []
        for sname in targets:
            ws = wb[sname]
            header, data = _rows_from_sheet(ws)
            if header is None:
                continue
            col_map = _header_map(header)
            for ri, row in enumerate(data, start=2):
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue
                rec = {col_map[j]: row[j] for j in col_map if j < len(row)}

                direction = _io_to_direction(rec.get("io_type"))
                if direction is None:
                    errors.append({"sheet": sname, "row": ri, "message": "매출/매입 유형 인식 실패"})
                    continue

                issue_date = _parse_date(rec.get("write_date")) or _parse_date(rec.get("issue_date_raw"))
                if not issue_date:
                    errors.append({"sheet": sname, "row": ri, "message": "작성일자 누락"})
                    continue

                amount = _dec(rec.get("amount"))
                vat = _dec(rec.get("vat"))
                total = _dec(rec.get("total"))
                if total == 0:
                    total = amount + vat
                if amount == 0 and total == 0:
                    errors.append({"sheet": sname, "row": ri, "message": "공급가액/합계 모두 0"})
                    continue

                due = _parse_date(rec.get("due_date"))
                parsed.append({
                    "entity_biz_no": _norm_biz(rec.get("entity_biz_no")),
                    "direction": direction,
                    "issue_date": issue_date.isoformat(),
                    "counterparty": (str(rec.get("counterparty") or "").strip() or "(거래처 미상)")[:200],
                    "counterparty_biz_no": _norm_biz(rec.get("counterparty_biz_no")),
                    "description": (str(rec.get("description") or "").strip() or None),
                    "amount": float(amount),
                    "vat": float(vat),
                    "total": float(total),
                    "tax_type": (str(rec.get("tax_type") or "").strip() or None),
                    "due_date": due.isoformat() if due else None,
                    "amend": (str(rec.get("amend") or "").strip() or None),
                    "raw": {k: (str(v) if v is not None else None) for k, v in rec.items()},
                })

        sales = sum(1 for p in parsed if p["direction"] == "sales")
        purchase = sum(1 for p in parsed if p["direction"] == "purchase")
        return {
            "parsed": parsed,
            "errors": errors,
            "stats": {
                "total": len(parsed) + len(errors),
                "valid": len(parsed),
                "errors": len(errors),
                "sales": sales,
                "purchase": purchase,
            },
        }
    finally:
        wb.close()

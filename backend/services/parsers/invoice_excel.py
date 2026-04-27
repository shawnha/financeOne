"""세금계산서 Excel 파서 (홈택스 / 회계법인 발행 양식).

표준 홈택스 컬럼 자동 매핑:
- 작성일자 → issue_date
- 승인번호 → document_no
- 공급자 등록번호 / 사업자번호 → seller_biz_no
- 공급받는자 등록번호 → buyer_biz_no
- 상호 (공급자/공급받는자 별도 컬럼) → 거래처
- 공급가액 → amount
- 세액 → vat
- 합계금액 → total
- 품목 / 거래내용 / 비고 → description
- 결제일자 / 결제예정일 → due_date

direction 판별 (our_biz_no 인자 기반):
- our_biz_no 가 공급자 측 → sales (우리가 발행, 받을 돈)
- our_biz_no 가 공급받는자 측 → purchase (우리가 수취, 줄 돈)
- 둘 다 일치 안 함 → 'unknown' 으로 표시 (사용자 수동 결정 필요)

비표준 컬럼명도 키워드 부분일치로 fuzzy 매핑 — header_to_field 함수 참조.
"""

from __future__ import annotations

import io
from datetime import date
from decimal import Decimal
from typing import Optional

import openpyxl
import xlrd


# ── 헤더 키워드 매핑 ───────────────────────────────────────────────────


# 키: 우리 표준 필드 / 값: 가능한 헤더 substring 들 (소문자 비교)
HEADER_KEYWORDS = {
    "issue_date": ["작성일자", "발행일자", "발행일", "거래일자", "issue_date"],
    "document_no": ["승인번호", "문서번호", "관리번호", "approval", "document_no"],
    "seller_biz_no": ["공급자.*등록번호", "공급자사업자", "발행자.*사업자"],
    "buyer_biz_no": ["공급받는자.*등록번호", "공급받는자사업자", "매입자.*사업자"],
    "seller_name": ["공급자.*상호", "공급자.*명", "발행자.*상호"],
    "buyer_name": ["공급받는자.*상호", "공급받는자.*명", "매입자.*상호"],
    "amount": ["공급가액", "공급금액"],
    "vat": ["세액", "부가세"],
    "total": ["합계금액", "합계", "총액"],
    "description": ["품목", "거래내용", "비고", "내용"],
    "due_date": ["결제일자", "결제예정일", "지급일자", "due_date"],
}


def header_to_field(header: str) -> Optional[str]:
    """헤더 문자열 → 표준 필드명. 매칭 안 되면 None."""
    import re
    if not header:
        return None
    h = str(header).strip().lower().replace(" ", "")
    for field, keywords in HEADER_KEYWORDS.items():
        for kw in keywords:
            kw_clean = kw.lower().replace(" ", "")
            # 정규식 패턴 (.* 포함) 또는 단순 substring
            if ".*" in kw_clean:
                if re.search(kw_clean, h):
                    return field
            elif kw_clean in h:
                return field
    return None


def _normalize_biz_no(s: str) -> str:
    """사업자번호 정규화 (하이픈/공백 제거, 10자리)."""
    if not s:
        return ""
    return "".join(c for c in str(s) if c.isdigit())


def _parse_date(v) -> Optional[date]:
    """다양한 형식 → date."""
    if v is None or v == "":
        return None
    if isinstance(v, date):
        return v
    s = str(v).strip()
    if not s or s.lower() in ("none", "nan"):
        return None
    # 8 digit YYYYMMDD
    digits = "".join(c for c in s if c.isdigit())
    if len(digits) == 8:
        try:
            return date(int(digits[:4]), int(digits[4:6]), int(digits[6:8]))
        except ValueError:
            pass
    # YYYY-MM-DD or YYYY/MM/DD or YYYY.MM.DD
    for sep in ("-", "/", "."):
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) >= 3:
                try:
                    y, m, d = int(parts[0]), int(parts[1]), int(parts[2])
                    if y < 100:
                        y += 2000
                    return date(y, m, d)
                except (ValueError, IndexError):
                    continue
    return None


def _parse_decimal(v) -> Decimal:
    if v is None or v == "":
        return Decimal("0")
    s = str(v).replace(",", "").replace("₩", "").strip()
    if not s or s.lower() in ("none", "nan", "-"):
        return Decimal("0")
    try:
        return Decimal(s)
    except Exception:
        return Decimal("0")


def _open_workbook_rows(file_bytes: bytes, filename: str) -> list[list]:
    """xls/xlsx 자동 감지 → 첫 sheet 의 모든 row 를 list[list] 로 반환."""
    name = filename.lower()
    if name.endswith(".xls"):
        wb = xlrd.open_workbook(file_contents=file_bytes)
        sheet = wb.sheet_by_index(0)
        rows = []
        for r in range(sheet.nrows):
            rows.append([sheet.cell_value(r, c) for c in range(sheet.ncols)])
        return rows
    # xlsx 또는 unknown → openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    sheet = wb[wb.sheetnames[0]]
    rows = []
    for row in sheet.iter_rows(values_only=True):
        rows.append(list(row))
    wb.close()
    return rows


def _detect_header_row(rows: list[list]) -> tuple[int, dict[int, str]]:
    """헤더 행 자동 탐지 + 컬럼 매핑.

    score = 매핑된 표준 필드 개수. score 가 가장 큰 row 를 헤더로 선택.
    Returns: (header_row_index, {col_index: standard_field}).
    """
    best_idx = 0
    best_map: dict[int, str] = {}
    for i, row in enumerate(rows[:30]):  # 첫 30행 스캔
        col_map = {}
        for j, cell in enumerate(row):
            if cell is None:
                continue
            field = header_to_field(str(cell))
            if field:
                col_map[j] = field
        if len(col_map) > len(best_map):
            best_idx = i
            best_map = col_map
    return best_idx, best_map


# ── Public ────────────────────────────────────────────────────────────


def parse_invoice_excel(
    file_bytes: bytes,
    filename: str,
    our_biz_no: Optional[str] = None,
) -> dict:
    """세금계산서 Excel → 파싱된 invoice 후보 + 검증 결과.

    Args:
        file_bytes: Excel binary.
        filename: 확장자 추론용.
        our_biz_no: 우리 사업자번호. direction 자동 판별에 사용. 없으면 모두 'unknown'.

    Returns:
        {
            "parsed": [<invoice dict>, ...],
            "errors": [{"row": <1-based>, "message": ...}, ...],
            "header_row": <int>,
            "column_map": {col_idx: field_name, ...},
            "stats": {"total": N, "valid": M, "errors": K, "unknown_direction": L},
        }

        invoice dict 필드:
        direction (sales/purchase/unknown), counterparty, counterparty_biz_no,
        issue_date (str ISO), due_date, document_no, amount, vat, total,
        description, raw (원본 row dict).
    """
    rows = _open_workbook_rows(file_bytes, filename)
    if not rows:
        return {"parsed": [], "errors": [], "header_row": 0, "column_map": {}, "stats": {"total": 0, "valid": 0, "errors": 0, "unknown_direction": 0}}

    header_idx, col_map = _detect_header_row(rows)
    if not col_map:
        return {
            "parsed": [], "errors": [{"row": 0, "message": "표준 컬럼을 인식하지 못했습니다 (작성일자/공급가액 등 누락)."}],
            "header_row": header_idx, "column_map": {},
            "stats": {"total": 0, "valid": 0, "errors": 1, "unknown_direction": 0},
        }

    our_biz_clean = _normalize_biz_no(our_biz_no) if our_biz_no else ""
    parsed = []
    errors = []
    unknown_dir = 0

    for row_i, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
        # 빈 row skip
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue

        record_raw = {col_map[j]: row[j] for j in col_map if j < len(row)}

        # 필수 필드 검증
        issue_date = _parse_date(record_raw.get("issue_date"))
        if not issue_date:
            errors.append({"row": row_i, "message": "작성일자 누락 또는 파싱 실패"})
            continue

        amount = _parse_decimal(record_raw.get("amount"))
        vat = _parse_decimal(record_raw.get("vat"))
        total = _parse_decimal(record_raw.get("total"))
        if total == 0:
            total = amount + vat
        if amount == 0 and total == 0:
            errors.append({"row": row_i, "message": "공급가액/합계 모두 0"})
            continue

        seller_biz = _normalize_biz_no(str(record_raw.get("seller_biz_no", "") or ""))
        buyer_biz = _normalize_biz_no(str(record_raw.get("buyer_biz_no", "") or ""))
        seller_name = str(record_raw.get("seller_name", "") or "").strip()
        buyer_name = str(record_raw.get("buyer_name", "") or "").strip()

        # direction 자동 판별
        if our_biz_clean and seller_biz == our_biz_clean:
            direction = "sales"
            counterparty = buyer_name or "(거래처 미상)"
            counterparty_biz_no = buyer_biz or None
        elif our_biz_clean and buyer_biz == our_biz_clean:
            direction = "purchase"
            counterparty = seller_name or "(거래처 미상)"
            counterparty_biz_no = seller_biz or None
        else:
            direction = "unknown"
            counterparty = (seller_name or buyer_name or "(거래처 미상)")
            counterparty_biz_no = seller_biz or buyer_biz or None
            unknown_dir += 1

        parsed.append({
            "direction": direction,
            "counterparty": counterparty[:200],
            "counterparty_biz_no": counterparty_biz_no,
            "issue_date": issue_date.isoformat(),
            "due_date": _parse_date(record_raw.get("due_date")).isoformat() if _parse_date(record_raw.get("due_date")) else None,
            "document_no": str(record_raw.get("document_no", "") or "").strip() or None,
            "amount": float(amount),
            "vat": float(vat),
            "total": float(total),
            "description": str(record_raw.get("description", "") or "").strip()[:500] or None,
            "row_number": row_i,
            "raw": {k: (str(v) if v is not None else None) for k, v in record_raw.items()},
        })

    return {
        "parsed": parsed,
        "errors": errors,
        "header_row": header_idx,
        "column_map": {str(k): v for k, v in col_map.items()},
        "stats": {
            "total": len(parsed) + len(errors),
            "valid": len(parsed),
            "errors": len(errors),
            "unknown_direction": unknown_dir,
        },
    }

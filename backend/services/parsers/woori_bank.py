"""우리은행 .xlsx parser."""

import io
import logging
import re
import openpyxl

logger = logging.getLogger(__name__)
import openpyxl.styles.colors as _colors
from datetime import date, datetime
from typing import Optional

from .base import BaseParser, ParsedTransaction, ParseResult
from .utils import parse_amount, parse_date

# Patch openpyxl aRGB regex to accept both 6-char (#RRGGBB) and 8-char (AARRGGBB)
# hex colors.  Korean bank exports frequently emit short-form hex values that
# openpyxl 3.x rejects by default.
_colors.aRGB_REGEX = re.compile(r"^#?([A-Fa-f0-9]{8}|[A-Fa-f0-9]{6})$")


class WooriBankParser(BaseParser):
    """Parse 우리은행 거래내역조회 .xlsx files."""

    def detect(self, file_bytes: bytes, filename: str) -> bool:
        if not filename.lower().endswith(".xlsx"):
            return False
        try:
            wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            sheet = wb.active
            # Row 1 (1-indexed) should contain "우리은행 거래내역조회"
            cell_val = str(sheet.cell(row=1, column=1).value or "").strip()
            wb.close()
            return "우리은행" in cell_val and "거래내역" in cell_val
        except Exception:
            return False

    def parse(self, file_bytes: bytes, filename: str) -> list[ParsedTransaction]:
        return self.parse_with_balance(file_bytes, filename).transactions

    def parse_with_balance(self, file_bytes: bytes, filename: str) -> ParseResult:
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
        sheet = wb.active

        results: list[ParsedTransaction] = []
        first_balance: float | None = None
        last_balance: float | None = None
        last_amount: float = 0
        last_type: str = "out"
        last_date: date | None = None

        # 1-indexed: Row 4 = headers, data starts at row 5
        # Columns (1-indexed): B=2(거래일시), C=3(적요), D=4(기재내용), E=5(지급), F=6(입금), G=7(잔액)
        for row in sheet.iter_rows(min_row=5, values_only=False):
            try:
                # Column B (index 1): 거래일시
                date_cell = row[1].value
                if date_cell is None:
                    continue

                if isinstance(date_cell, datetime):
                    tx_date = date_cell.date()
                elif isinstance(date_cell, date):
                    tx_date = date_cell
                else:
                    date_str = str(date_cell).strip()
                    # Format: YYYY.MM.DD HH:MM:SS
                    tx_date = parse_date(date_str.split(" ")[0])
                    if tx_date is None:
                        continue

                # Column C (index 2): 적요
                memo = str(row[2].value or "").strip()

                # Column D (index 3): 기재내용
                description_detail = str(row[3].value or "").strip()

                # Column E (index 4): 지급
                withdrawal_raw = row[4].value
                withdrawal = parse_amount(str(withdrawal_raw)) if withdrawal_raw else None

                # Column F (index 5): 입금
                deposit_raw = row[5].value
                deposit = parse_amount(str(deposit_raw)) if deposit_raw else None

                # Determine type and amount
                if withdrawal and withdrawal > 0:
                    tx_type = "out"
                    amount = abs(withdrawal)
                elif deposit and deposit > 0:
                    tx_type = "in"
                    amount = abs(deposit)
                else:
                    continue

                # Column G (index 6): 거래 후 잔액
                balance_raw = row[6].value if len(row) > 6 else None
                if balance_raw is not None:
                    bal = parse_amount(str(balance_raw))
                    if bal is not None:
                        if first_balance is None:
                            first_balance = bal
                        last_balance = bal
                        last_amount = amount
                        last_type = tx_type
                        last_date = tx_date

                # 적요 = '체크우리' means check card transaction
                is_check_card = memo == "체크우리"

                counterparty = description_detail
                description = f"{memo} {description_detail}".strip()

                # 원본 데이터 보존
                raw = {
                    "번호": str(row[0].value or ""),
                    "거래일시": str(date_cell),
                    "적요": memo,
                    "기재내용": description_detail,
                    "찾으신금액": float(withdrawal) if withdrawal else 0,
                    "맡기신금액": float(deposit) if deposit else 0,
                    "거래후잔액": float(bal) if balance_raw is not None and bal is not None else None,
                }
                # 나머지 컬럼도 보존
                for ci in range(7, len(row)):
                    cell = row[ci]
                    if cell.value is not None:
                        raw[f"col_{ci}"] = str(cell.value)

                results.append(ParsedTransaction(
                    date=tx_date,
                    amount=amount,
                    currency="KRW",
                    type=tx_type,
                    description=description,
                    counterparty=counterparty,
                    source_type="woori_bank",
                    is_check_card=is_check_card,
                    raw_data=raw,
                    row_number=row[0].row,
                    balance_after=float(bal) if balance_raw is not None and bal is not None else None,
                ))
            except Exception as e:
                logger.warning("Parse row failed: %s", e)
                continue

        wb.close()
        # 우리은행 Excel은 최신순 (첫 행=최신, 마지막 행=가장 오래된 거래)
        # first_balance = 기말잔고 (첫 행 = 가장 최신 거래 후 잔액)
        # last_balance = 가장 오래된 거래의 거래 후 잔액
        # 기초잔고 = 가장 오래된 거래의 거래 전 잔액 (역산)
        opening = last_balance
        if opening is not None and last_type == "out":
            opening = last_balance + last_amount  # 지급이면 잔액에 다시 더함
        elif opening is not None and last_type == "in":
            opening = last_balance - last_amount  # 입금이면 잔액에서 뺌

        # 시간순(오래된→최신)으로 뒤집어서 반환 — DB INSERT 시 id 순서 = 시간순
        results.reverse()
        return ParseResult(
            transactions=results,
            opening_balance=opening,  # 가장 오래된 거래 전 잔액 = 기초잔고
            closing_balance=first_balance,  # 첫 행 = 가장 최신 = 기말
            balance_date=results[-1].date if results else None,
        )
